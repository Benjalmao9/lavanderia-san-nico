# ============================================================
#  Generación del reporte contable en Excel (.xlsx) con openpyxl.
#
#  Este módulo NO consulta la base de datos: recibe los datos ya resumidos (los
#  arma routers/reportes.py reutilizando sus consultas existentes) y devuelve los
#  bytes de un archivo .xlsx con 5 pestañas.
#
#  ── ¿POR QUÉ GRÁFICOS NATIVOS DE EXCEL Y NO IMÁGENES? ──────────────────────
#  Un gráfico NATIVO se guarda como datos dentro del .xlsx: Excel lo dibuja al
#  abrirlo, queda ligado a las celdas y el usuario puede interactuar con él
#  (pasar el mouse, cambiar colores, reusarlo en otra hoja, copiarlo a un
#  informe). Una imagen (PNG) sería una foto muerta: no se puede editar, se ve
#  borrosa al hacer zoom o imprimir, y no refleja los números si alguien ajusta
#  la tabla. Para un documento de contabilidad, lo nativo es más útil y
#  profesional. openpyxl.chart genera justamente eso (LineChart/BarChart/PieChart
#  referenciando celdas), sin depender de librerías de imágenes.
#
#  ── ¿POR QUÉ LAS FECHAS VAN SIEMPRE EN HORA DE MÉXICO (UTC-6)? ─────────────
#  La interfaz EN VIVO convierte las fechas a la zona del DISPOSITIVO de quien
#  mira (ver frontend/src/utils/formato.ts): así cada usuario ve su hora local.
#  Pero este Excel es un documento de contabilidad FIJO: una vez descargado, el
#  mismo archivo lo puede abrir el dueño en México, un contador en otra zona o
#  reenviarse por correo. Si dejáramos que Excel lo interpretara según la PC que
#  lo abre, las horas "cambiarían" de una computadora a otra: inaceptable en un
#  registro contable. Por eso convertimos explícitamente cada fecha (guardada en
#  UTC, ver tiempo.py) a hora de Ciudad de México ANTES de escribirla, y la
#  escribimos como hora "de pared" fija. México ya no aplica horario de verano
#  (desde 2022), así que un offset fijo de UTC-6 es correcto y estable.
# ============================================================

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# a_zona_mexico: la conversión UTC -> hora de México, COMPARTIDA con el cálculo
# de "hoy" del negocio en tiempo.py (antes esta zona horaria estaba duplicada
# acá; centralizarla evita que las dos copias se desalineen).
from tiempo import a_zona_mexico

# Formatos de celda (códigos de formato de Excel).
_FMT_MONEDA = '"$"#,##0.00'          # dinero: $1,234.50
_FMT_NUMERO = "#,##0.00"             # cantidades con decimales (kilos)
_FMT_ENTERO = "#,##0"                # conteos
_FMT_FECHA = "DD/MM/YYYY HH:MM"       # fecha legible con hora

# Estilos de encabezado: negrita blanca sobre azul acero (la marca del proyecto).
_FUENTE_ENCABEZADO = Font(bold=True, color="FFFFFF")
_RELLENO_ENCABEZADO = PatternFill("solid", fgColor="5B8FC7")  # #5b8fc7 (azul acero)
_ALINEAR_ENCABEZADO = Alignment(horizontal="center", vertical="center")

# Caracteres con los que una celda de TEXTO podría interpretarse como FÓRMULA al
# abrir el archivo (inyección de fórmulas de Excel / "CSV injection").
_PREFIJOS_PELIGROSOS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _texto_seguro(valor):
    """Sanea un valor de texto antes de escribirlo en una celda del .xlsx.

    Hace DOS cosas sobre los strings (los números/fechas se escriben como tales y
    no corren riesgo, así que se devuelven sin tocar):

    1) QUITA CARACTERES DE CONTROL ILEGALES en XLSX (\\x00-\\x08, \\x0b, \\x0c,
       \\x0e-\\x1f). Son necesarios porque los campos de texto que llenan los
       usuarios (cliente, notas...) no restringen el charset, y si uno trae, por
       ejemplo, un \\x1b o un \\x00, openpyxl lanza IllegalCharacterError al
       guardar y TODA la exportación de ese rango se caería con un 500. Los
       eliminamos con la expresión oficial de openpyxl (ILLEGAL_CHARACTERS_RE),
       que respeta el tab, el salto de línea y el retorno de carro (sí legales).

    2) NEUTRALIZA LA INYECCIÓN DE FÓRMULAS: si el texto empieza con '=', '+', '-',
       '@' o un carácter de control (tab/CR/LF), Excel/LibreOffice podría
       EJECUTARLO como fórmula al abrir el archivo (un cliente llamado '=SUM(...)'
       o '=cmd|...' es un vector de ataque, y openpyxl además lo guardaría como
       fórmula en vez de texto). Le anteponemos un apóstrofo, que fuerza a
       tratarlo como TEXTO literal.
    """
    if not isinstance(valor, str):
        return valor
    # 1) Depurar caracteres de control ilegales (evita el 500 al guardar).
    valor = ILLEGAL_CHARACTERS_RE.sub("", valor)
    # 2) Anti-inyección de fórmulas (sobre el valor ya depurado).
    if valor and valor[0] in _PREFIJOS_PELIGROSOS:
        return "'" + valor
    return valor


def _a_hora_mexico(dt):
    """Convierte un datetime naive-UTC (como se guarda en la BD) a hora de México.

    Devuelve un datetime naive con el "reloj de pared" de México (Excel no guarda
    zona horaria; guarda el valor tal cual y lo muestra igual en cualquier PC).
    None se mantiene como None (fecha ausente, p. ej. un pedido sin entregar).
    La conversión en sí vive en tiempo.py (compartida con hoy_mexico()).
    """
    if dt is None:
        return None
    return a_zona_mexico(dt)


def _estilar_encabezado(ws, num_columnas: int, fila: int = 1) -> None:
    """Pone en negrita/color la fila de encabezados (celdas 1..num_columnas)."""
    for col in range(1, num_columnas + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = _FUENTE_ENCABEZADO
        celda.fill = _RELLENO_ENCABEZADO
        celda.alignment = _ALINEAR_ENCABEZADO


def _ajustar_anchos(ws, minimo: int = 10, maximo: int = 45) -> None:
    """Ajusta el ancho de cada columna al contenido más largo (con topes).

    Estima el ancho por la longitud del texto de cada celda (para fechas usa el
    largo del formato, no el repr largo del datetime). Acota entre 'minimo' y
    'maximo' para que ninguna columna quede diminuta ni gigante.
    """
    anchos: dict[int, int] = {}
    for fila in ws.iter_rows():
        for celda in fila:
            if celda.value is None:
                continue
            if isinstance(celda.value, datetime):
                largo = len("DD/MM/YYYY HH:MM")
            else:
                largo = len(str(celda.value))
            anchos[celda.column] = max(anchos.get(celda.column, 0), largo)
    for col, largo in anchos.items():
        ancho = min(max(largo + 2, minimo), maximo)
        ws.column_dimensions[get_column_letter(col)].width = ancho


def _escribir_tabla(ws, encabezados, filas, formatos=None) -> int:
    """Escribe una tabla (encabezados en la fila 1, datos debajo) y la estiliza.

    - encabezados: lista de títulos de columna.
    - filas: lista de tuplas/listas con los valores (mismo orden que encabezados).
    - formatos: dict {indice_columna_0based: codigo_de_formato} para dinero/fecha.
    Devuelve la cantidad de filas de datos escritas (para ubicar el gráfico).
    """
    ws.append(encabezados)
    _estilar_encabezado(ws, len(encabezados))
    for fila in filas:
        # _texto_seguro protege cada celda de texto contra inyección de fórmulas.
        ws.append([_texto_seguro(v) for v in fila])
    # Aplicamos formatos de número/fecha a las columnas indicadas (desde la fila 2).
    if formatos:
        for idx_col, fmt in formatos.items():
            for fila_num in range(2, len(filas) + 2):
                ws.cell(row=fila_num, column=idx_col + 1).number_format = fmt
    _ajustar_anchos(ws)
    return len(filas)


def _hoja_ingresos(wb, datos) -> None:
    """Pestaña 'Ingresos por periodo': tabla + gráfico de LÍNEA nativo."""
    ws = wb.active  # reutilizamos la hoja por defecto como primera pestaña
    ws.title = "Ingresos por periodo"
    filas = [(d["periodo"], float(d["ingresos"])) for d in datos]
    n = _escribir_tabla(
        ws, ["Periodo", "Ingresos"], filas, formatos={1: _FMT_MONEDA}
    )
    if n == 0:
        ws["A3"] = "Sin datos en el rango seleccionado."
        return
    chart = LineChart()
    chart.title = "Ingresos por periodo"
    chart.style = 12
    chart.y_axis.title = "Ingresos ($)"
    chart.x_axis.title = "Periodo"
    # delete=False fuerza que los ejes se dibujen (quirk conocido de openpyxl).
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # Serie de datos: incluye la fila 1 como NOMBRE de la serie (titles_from_data).
    datos_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    categorias = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(datos_ref, titles_from_data=True)
    chart.set_categories(categorias)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "D2")  # anclado a la derecha de la tabla


def _hoja_por_estado(wb, datos) -> None:
    """Pestaña 'Pedidos por estado': tabla + gráfico de PASTEL nativo."""
    ws = wb.create_sheet("Pedidos por estado")
    filas = [(d["estado"], d["cantidad"]) for d in datos]
    n = _escribir_tabla(ws, ["Estado", "Cantidad"], filas, formatos={1: _FMT_ENTERO})
    if n == 0:
        ws["A3"] = "Sin datos en el rango seleccionado."
        return
    chart = PieChart()
    chart.title = "Pedidos por estado"
    datos_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    categorias = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(datos_ref, titles_from_data=True)
    chart.set_categories(categorias)
    # Mostrar el valor en cada porción para que se lea sin pasar el mouse.
    chart.dataLabels = _etiquetas_valor()
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, "D2")


def _hoja_por_periodo(wb, datos) -> None:
    """Pestaña 'Pedidos por periodo': tabla + gráfico de BARRAS (verticales)."""
    ws = wb.create_sheet("Pedidos por periodo")
    filas = [(d["periodo"], d["cantidad"]) for d in datos]
    n = _escribir_tabla(ws, ["Periodo", "Cantidad"], filas, formatos={1: _FMT_ENTERO})
    if n == 0:
        ws["A3"] = "Sin datos en el rango seleccionado."
        return
    chart = BarChart()
    chart.type = "col"  # columnas verticales
    chart.title = "Pedidos por periodo"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Periodo"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend = None
    datos_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    categorias = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(datos_ref, titles_from_data=True)
    chart.set_categories(categorias)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "D2")


def _hoja_por_empleado(wb, datos) -> None:
    """Pestaña 'Pedidos por empleado': tabla + gráfico de BARRAS HORIZONTALES."""
    ws = wb.create_sheet("Pedidos por empleado")
    filas = [(d["nombre_completo"], d["cantidad"]) for d in datos]
    n = _escribir_tabla(ws, ["Empleado", "Cantidad"], filas, formatos={1: _FMT_ENTERO})
    if n == 0:
        ws["A3"] = "Sin datos en el rango seleccionado."
        return
    chart = BarChart()
    chart.type = "bar"  # barras HORIZONTALES (bar = horizontal; col = vertical)
    chart.title = "Pedidos por empleado"
    chart.x_axis.title = "Cantidad"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend = None
    datos_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    categorias = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(datos_ref, titles_from_data=True)
    chart.set_categories(categorias)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "D2")


def _hoja_detalle(wb, detalle) -> None:
    """Pestaña 'Detalle de pedidos': listado línea por línea (sin gráfico)."""
    ws = wb.create_sheet("Detalle de pedidos")
    encabezados = [
        "ID", "Cliente", "Teléfono", "Kilos", "Precio por kilo", "Total",
        "Estado", "Fecha de recepción", "Fecha de entrega", "Empleado", "Notas",
    ]
    filas = []
    for p in detalle:
        filas.append((
            p["id"],
            p["cliente"],
            p["telefono"],
            float(p["kilos"]) if p["kilos"] is not None else None,
            float(p["precio_por_kilo"]) if p["precio_por_kilo"] is not None else None,
            float(p["total"]) if p["total"] is not None else None,
            p["estado"],
            _a_hora_mexico(p["fecha_recepcion"]),   # UTC -> hora de México
            _a_hora_mexico(p["fecha_entrega"]),
            p["empleado"],
            p["notas"],
        ))
    # Formatos: kilos (3), precio (4) y total (5) money/numero; fechas (7,8).
    formatos = {
        3: _FMT_NUMERO,
        4: _FMT_MONEDA,
        5: _FMT_MONEDA,
        7: _FMT_FECHA,
        8: _FMT_FECHA,
    }
    n = _escribir_tabla(ws, encabezados, filas, formatos=formatos)
    if n == 0:
        # Aun sin pedidos, dejamos los encabezados y una nota clara.
        ws.append(["Sin pedidos en el rango seleccionado."])


def _etiquetas_valor():
    """DataLabels que muestran el VALOR en el gráfico de pastel."""
    from openpyxl.chart.label import DataLabelList

    etiquetas = DataLabelList()
    etiquetas.showVal = True
    return etiquetas


def construir_reporte_excel(
    ingresos, por_periodo, por_estado, por_empleado, detalle
) -> bytes:
    """Arma el .xlsx completo (5 pestañas) y devuelve sus bytes.

    Recibe los datos YA consultados (reportes.py los obtiene reutilizando sus
    consultas existentes). El orden de las pestañas es el pedido:
      1) Ingresos por periodo   2) Pedidos por estado
      3) Pedidos por periodo    4) Pedidos por empleado
      5) Detalle de pedidos
    """
    wb = Workbook()
    _hoja_ingresos(wb, ingresos)        # usa la hoja por defecto (primera)
    _hoja_por_estado(wb, por_estado)
    _hoja_por_periodo(wb, por_periodo)
    _hoja_por_empleado(wb, por_empleado)
    _hoja_detalle(wb, detalle)

    # Guardamos en memoria (BytesIO) y devolvemos los bytes: el endpoint los manda
    # como descarga sin tocar el disco.
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
